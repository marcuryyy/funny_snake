from contextlib import asynccontextmanager
from fastapi import Depends, FastAPI, HTTPException, Query, Response, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from datetime import date, datetime
import csv
import io
from typing import Annotated, Optional, List, Union

from fastapi.security import HTTPBasic, HTTPBasicCredentials
from mail_fetch import fetch_emails
from mail_sending import send_email
import asyncpg
import bcrypt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os
import uvicorn
from pydantic_models import (
    AddNewRow,
    RequestCreate,
    RequestResponse,
    FetchedMailsResponse,
    EmailRequest,
)
from utils import parse_date_string, date_to_str

POSTGRES_DB_NAME = os.getenv("POSTGRES_DB", "postgres")
POSTGRES_DB_USER = os.getenv("POSTGRES_USER", "postgres")
POSTGRES_DB_PASS = os.getenv("POSTGRES_PASSWORD", "postgres")
POSTGRES_HOSTNAME = os.getenv("POSTGRES_HOSTNAME", "postgres")
POSTGRES_PORT = os.getenv("POSTGRES_PORT", "5432")



async def get_filtered_requests(
    db_pool,
    full_name: Optional[str],
    object_name: Optional[str],
    phone: Optional[str],
    email: Optional[str],
    emotion: Optional[str],
    issue: Optional[str],
    date_from: Optional[str],
    date_to: Optional[str],
    task_status: Optional[str],
    limit: Optional[int] = None,
    offset: Optional[int] = None,
) -> List[RequestResponse]:
    """Возвращает отфильтрованные запросы из БД"""
    if not db_pool:
        raise HTTPException(status_code=503, detail="DB not ready")

    parsed_date_from = parse_date_string(date_from) if date_from else None
    parsed_date_to = parse_date_string(date_to) if date_to else None

    params = [
        f"%{full_name}%" if full_name else None,
        f"%{object_name}%" if object_name else None,
        phone,
        f"%{email}%" if email else None,
        emotion,
        f"%{issue}%" if issue else None,
        parsed_date_from,
        parsed_date_to,
        task_status.upper() if task_status else None,
    ]

    async with db_pool.acquire() as conn:
        base_query = """
            SELECT * FROM requests
            WHERE ($1::text IS NULL OR LOWER(full_name) LIKE LOWER($1))
              AND ($2::text IS NULL OR LOWER(object_name) LIKE LOWER($2))
              AND ($3::text IS NULL OR phone ILIKE $3)
              AND ($4::text IS NULL OR LOWER(email) LIKE LOWER($4))
              AND ($5::text IS NULL OR emotion = $5)
              AND ($6::text IS NULL OR LOWER(question_summary) LIKE LOWER($6))
              AND ($7::date IS NULL OR req_date >= $7)
              AND ($8::date IS NULL OR req_date <= $8)
              AND ($9::text IS NULL OR task_status = $9::task_statuses)
        """

        if limit is not None and offset is not None:
            base_query += " ORDER BY request_id ASC LIMIT $10 OFFSET $11"
            params.extend([limit, offset])
        else:
            base_query += " ORDER BY request_id ASC"
        rows = await conn.fetch(base_query, *params)

        result = []
        for row in rows:
            print(row["llm_answer"])
            result.append(
                RequestResponse(
                    id=row["request_id"],
                    date=date_to_str(row["req_date"]),
                    fullName=row["full_name"] or "",
                    object=row["object_name"] or "",
                    phone=row["phone"] or "",
                    email=row["email"] or "",
                    factoryNumber=row["factory_number"] or "",
                    deviceType=row["device_type"] or "",
                    emotion=row["emotion"],
                    issue=row["question_summary"] or "",
                    llm_answer=row["llm_answer"] or "",
                    task_status=row["task_status"] or "OPEN",
                    message_id=row["message_id"] or "",
                )
            )
        return result


@asynccontextmanager
async def lifespan(app: FastAPI):
    try:
        pool = await asyncpg.create_pool(
            user=POSTGRES_DB_USER,
            password=POSTGRES_DB_PASS,
            database=POSTGRES_DB_NAME,
            host=POSTGRES_HOSTNAME,
            port=POSTGRES_PORT,
            min_size=2,
            max_size=10,
        )
        app.state.db_pool = pool

    except Exception as e:
        raise e

    yield

    await app.state.db_pool.close()


app = FastAPI(lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.get("/api/requests", response_model=List[RequestResponse])
async def get_requests(
    page: int = Query(1, ge=1),
    limit: int = Query(10, ge=1, le=100),
    full_name: Optional[str] = Query(None),
    object_name: Optional[str] = Query(None),
    phone: Optional[str] = Query(None),
    email: Optional[str] = Query(None),
    emotion: Optional[str] = Query(None),
    issue: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    task_status: Optional[str] = Query(None),
):
    """Получить список запросов с фильтрами и пагинацией"""
    offset = (page - 1) * limit

    result = await get_filtered_requests(
        db_pool=app.state.db_pool,
        full_name=full_name,
        object_name=object_name,
        phone=phone,
        email=email,
        emotion=emotion,
        issue=issue,
        date_from=date_from,
        date_to=date_to,
        task_status=task_status,
        limit=limit,
        offset=offset,
    )
    return result


@app.post("/api/requests", response_model=AddNewRow)
async def create_request(request_data: RequestCreate):
    db_pool = app.state.db_pool
    if not db_pool:
        raise HTTPException(status_code=503, detail="DB not ready")

    raw_date = request_data.date
    parsed_date = parse_date_string(raw_date)

    async with db_pool.acquire() as conn:
        query = """
            INSERT INTO requests
            (req_date, full_name, object_name, phone, email, factory_number, device_type, emotion, question_summary, llm_answer, task_status, message_id)
            VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12)
            RETURNING request_id
        """
        row = await conn.fetchrow(
            query,
            parsed_date,
            request_data.fullName,
            request_data.object,
            request_data.phone,
            request_data.email,
            request_data.factoryNumber,
            request_data.deviceType,
            request_data.emotion,
            request_data.issue,
            request_data.llm_answer,
            request_data.task_status or "OPEN",
            request_data.message_id or "",
        )

        return AddNewRow(id=row["request_id"])


@app.get("/api/fetchMails", response_model=List[FetchedMailsResponse])
async def get_mails():
    msgs = fetch_emails(limit=10, save_attachments_dir="attachments")
    return msgs


@app.post("/api/sendMail")
async def send_mail_endpoint(request: EmailRequest):
    success = await send_email(
        to_emails=request.to_emails,
        subject=request.subject,
        body=request.body,
        html_body=request.html_body,
        from_email=request.from_email,
        message_id=request.message_id,
        reply_to_thread=True,
    )

    if success:
        return {
            "status": "success",
            "message": f"Email отправлен на {len(request.to_emails)} адресов",
            "recipients": request.to_emails,
        }
    else:
        raise HTTPException(status_code=500, detail="Ошибка отправки email")


@app.get("/api/getCsv")
async def get_table_csv(
    full_name: Optional[str] = Query(None),
    object_name: Optional[str] = Query(None),
    phone: Optional[str] = Query(None),
    email: Optional[str] = Query(None),
    emotion: Optional[str] = Query(None),
    issue: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    task_status: Optional[str] = Query(None),
):
    """Экспорт отфильтрованных запросов в CSV"""
    db_pool = app.state.db_pool
    if not db_pool:
        raise HTTPException(status_code=503, detail="DB not ready")

    async def generate_csv():
        async with db_pool.acquire() as conn:
            headers = list(RequestResponse.model_fields.keys())

            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(headers)
            yield output.getvalue()

            batch_size = 5000
            offset = 0

            while True:
                batch = await get_filtered_requests(
                    db_pool=db_pool,
                    full_name=full_name,
                    object_name=object_name,
                    phone=phone,
                    email=email,
                    emotion=emotion,
                    issue=issue,
                    date_from=date_from,
                    date_to=date_to,
                    task_status=task_status,
                    limit=batch_size,
                    offset=offset,
                )

                if not batch:
                    break

                csv_rows = []
                for row in batch:
                    row_dict = row.dict() if hasattr(row, "dict") else dict(row)
                    csv_row = [str(row_dict.get(h, "")) for h in headers]
                    csv_rows.append(csv_row)

                output = io.StringIO()
                writer = csv.writer(output)
                writer.writerows(csv_rows)
                yield output.getvalue()

                output.close()

                offset += batch_size

    filename = f"requests_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

    return StreamingResponse(
        generate_csv(),
        media_type="text/csv",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Type": "text/csv; charset=utf-8-sig",
        },
    )


@app.get("/api/getExcel")
async def get_table_excel(
    full_name: Optional[str] = Query(None),
    object_name: Optional[str] = Query(None),
    phone: Optional[str] = Query(None),
    email: Optional[str] = Query(None),
    emotion: Optional[str] = Query(None),
    issue: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    task_status: Optional[str] = Query(None),
):
    """Экспорт отфильтрованных запросов в Excel"""
    db_pool = app.state.db_pool
    if not db_pool:
        raise HTTPException(status_code=503, detail="DB not ready")

    wb = Workbook()
    ws = wb.active
    ws.title = "Requests"

    headers = list(RequestResponse.model_fields.keys())

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    batch_size = 1000
    offset = 0
    row_num = 2

    while True:
        batch = await get_filtered_requests(
            db_pool=db_pool,
            full_name=full_name,
            object_name=object_name,
            phone=phone,
            email=email,
            emotion=emotion,
            issue=issue,
            date_from=date_from,
            date_to=date_to,
            task_status=task_status,
            limit=batch_size,
            offset=offset,
        )

        if not batch:
            break

        for row in batch:
            if hasattr(row, "dict"):
                row_dict = row.dict()
            elif hasattr(row, "__dict__"):
                row_dict = row.__dict__
            else:
                row_dict = dict(row)

            for col, header in enumerate(headers, 1):
                value = row_dict.get(header, "")
                ws.cell(
                    row=row_num,
                    column=col,
                    value=str(value) if value is not None else "",
                )
            row_num += 1

        offset += batch_size

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    file_content = output.getvalue()
    output.close()

    filename = f"requests_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return Response(
        content=file_content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )

security = HTTPBasic()

async def get_db_pool():
    """Получаем пул соединений из app.state"""
    if not app.state.db_pool:
        raise HTTPException(status_code=503, detail="Database not ready")
    return app.state.db_pool

async def get_current_username(
    credentials: Annotated[HTTPBasicCredentials, Depends(security)],
    db_pool=Depends(get_db_pool)
):
    """Проверяет пользователя в PostgreSQL"""
    
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT id, login, password_hash FROM users WHERE login = $1 AND is_active = true",
            credentials.username
        )
        
        if not row:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Неверный логин или пароль",
                headers={"WWW-Authenticate": "Basic"},
            )
        
        print("Пользователь найден")

        current_password_bytes = credentials.password.encode("utf8")
        stored_password_bytes = row['password_hash'].encode("utf8")
        
        is_correct_password = bcrypt.checkpw(
            current_password_bytes, 
            stored_password_bytes
        )
        
        if not is_correct_password:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Неверный логин или пароль",
                headers={"WWW-Authenticate": "Basic"},
            )
        
        return row['login']

@app.get("/users/me")
def read_current_user(username: Annotated[str, Depends(get_current_username)]):
    return {"username": username}

if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000)
